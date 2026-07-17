from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

FILE_NAME = "vehicle_data.xlsx"


# Create Excel file if it doesn't exist
def create_database():

    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicle Data"

        ws.append([
            "Timestamp",
            "VIN Number",
            "Model",
            "Fuel Type",
            "Issues"
        ])

        wb.save(FILE_NAME)


# Save Record
def save_record(vin, model, fuel, issues):

    create_database()

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    timestamp = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    ws.append([
        timestamp,
        vin,
        model,
        fuel,
        issues
    ])

    wb.save(FILE_NAME)
    wb.close()


# Start Work
def start_work():

    while True:

        print("\n===== START WORK =====")

        vin = int(input("Enter VIN Number : "))

        if vin == 0 :
            break
        print("\nSelect Model")
        print("1. Harrier")
        print("2. Safari")

        model_choice = input("Enter Choice : ")

        if model_choice == "1":
            model = "Harrier"
        else:
            model = "Safari"

        print("\nSelect Fuel Type")
        print("1. Diesel")
        print("2. Petrol")
        print("3. EV")

        fuel_choice = input("Enter Choice : ")

        if fuel_choice == "1":
            fuel = "Diesel"
            issue_list = [
                "RTB RH",
                "RTB LH",
                "Shockupser RH",
                "Shockupser LH",
                "Subframe RH Front",
		"Subframe RH Rear",
                "Subframe LH Front",
		"Subframe LH rear"
            ]

        elif fuel_choice == "2":
            fuel = "Petrol"
            issue_list = [
                "RTB RH",
                "RTB LH",
                 "Shockupser RH",
                "Shockupser LH",
                "Subframe RH Front",
                "Subframe RH Rear",
                "Subframe LH Front",
                "Subframe LH rear"
            ]

        else:
            fuel = "EV"
            issue_list = [
                "Moulting RH Front",
                "Moulting RH Rear",
                "Longi RH",
                "Longi LH",
                "Damper RH",
                "Damper LH",
                "Subframe RH Front",
                "Subframe RH Rear",
                "Subframe LH Front",
                "Subframe LH rear"
            ]

        print("\nSelect Issues (Example: 1,3,5)\n")

        for i, issue in enumerate(issue_list, 1):
            print(f"{i}. {issue}")

        selected = input("\nEnter Issue Numbers : ")

        selected_issues = []

        for num in selected.split(","):
            if num.strip().isdigit():
                index = int(num)
                if 1 <= index <= len(issue_list):
                    selected_issues.append(issue_list[index - 1])

        issues = ", ".join(selected_issues)

        save_record(vin, model, fuel, issues)

        print("\n✅ Record Saved Successfully.")

        print("\n Enter VIN 000 TO Exit or back to main manue")
